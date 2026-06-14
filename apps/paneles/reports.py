"""Reportes server-side con exportacion Excel y PDF."""

from io import BytesIO

from django.conf import settings
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import cm, mm
from reportlab.platypus import Flowable, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.cuentas.models import Usuario
from apps.academico.models import Aprendiz, Ficha, Instructor, Trimestre
from apps.proyectos.models import Proyecto

REPORT_LABELS = {
    'usuarios': 'Usuarios',
    'aprendices': 'Aprendices',
    'instructores': 'Instructores',
    'fichas': 'Fichas',
    'trimestres': 'Trimestres',
}

C_NAVY = colors.HexColor('#0A1628')
C_BLUE1 = colors.HexColor('#1E3A8A')
C_BLUE2 = colors.HexColor('#2563EB')
C_BLUE3 = colors.HexColor('#60A5FA')
C_ACCENT = colors.HexColor('#06B6D4')
C_ACTIVO = colors.HexColor('#10B981')
C_INACTIVO = colors.HexColor('#F43F5E')
C_AMBER = colors.HexColor('#F59E0B')
C_VIOLET = colors.HexColor('#7C3AED')
C_BG = colors.HexColor('#F0F4FF')
C_LIGHT = colors.HexColor('#EAF2FF')
C_BORDER = colors.HexColor('#BFDBFE')
C_ROW_ALT = colors.HexColor('#F4F8FF')
C_HEADER_ROW = colors.HexColor('#1D4ED8')


def _set_alpha(canvas, fill=None, stroke=None):
    if fill is not None and hasattr(canvas, 'setFillAlpha'):
        canvas.setFillAlpha(fill)
    if stroke is not None and hasattr(canvas, 'setStrokeAlpha'):
        canvas.setStrokeAlpha(stroke)


class SectionBar(Flowable):
    def __init__(self, text, width=15.9 * cm, height=16):
        super().__init__()
        self.text = text
        self.width = width
        self.height = height

    def draw(self):
        self.canv.setFillColor(C_LIGHT)
        self.canv.roundRect(0, 0, self.width, self.height, 4, stroke=0, fill=1)
        self.canv.setFillColor(C_BLUE2)
        self.canv.roundRect(0, 0, 5, self.height, 2, stroke=0, fill=1)
        self.canv.setFillColor(C_BLUE1)
        self.canv.setFont('Helvetica-Bold', 8)
        self.canv.drawString(13, 5, self.text.upper())


class AccentLine(Flowable):
    def __init__(self, width=15.9 * cm, height=8):
        super().__init__()
        self.width = width
        self.height = height

    def draw(self):
        y = self.height / 2
        self.canv.setLineWidth(2)
        self.canv.setStrokeColor(C_BLUE2)
        self.canv.line(0, y, self.width * 0.30, y)
        self.canv.setStrokeColor(C_ACCENT)
        self.canv.line(self.width * 0.30, y, self.width * 0.40, y)
        self.canv.setStrokeColor(C_BORDER)
        self.canv.line(self.width * 0.40, y, self.width, y)


class StatCard(Flowable):
    def __init__(self, value, label, accent=C_ACCENT, width=3.75 * cm, height=1.72 * cm):
        super().__init__()
        self.value = value
        self.label = label
        self.accent = accent
        self.width = width
        self.height = height

    def draw(self):
        self.canv.setFillColor(colors.HexColor('#D8E2F1'))
        self.canv.roundRect(2, -2, self.width, self.height, 5, stroke=0, fill=1)
        self.canv.setFillColor(C_NAVY)
        self.canv.roundRect(0, 0, self.width, self.height, 5, stroke=0, fill=1)
        self.canv.setFillColor(self.accent)
        self.canv.roundRect(0, self.height - 6, self.width, 6, 3, stroke=0, fill=1)
        self.canv.setFillColor(self.accent)
        self.canv.setFont('Helvetica-Bold', 17)
        self.canv.drawCentredString(self.width / 2, 23, str(self.value))
        self.canv.setFillColor(colors.HexColor('#DDE7F7'))
        self.canv.setFont('Helvetica-Bold', 5.8)
        self.canv.drawCentredString(self.width / 2, 9, self.label.upper())


def _logo_path():
    pdf_logo = settings.BASE_DIR / 'frontend_assets' / 'img' / 'logo3_pdf.png'
    if pdf_logo.exists():
        return pdf_logo
    pdf_logo = settings.BASE_DIR / 'frontend_assets' / 'img' / 'logo3_pdf.jpg'
    if pdf_logo.exists():
        return pdf_logo
    return settings.BASE_DIR / 'frontend_assets' / 'img' / 'logo3.png'


def _draw_sidebar(canvas, doc):
    page_w, page_h = doc.pagesize
    sidebar_w = 5 * mm
    canvas.saveState()
    canvas.setFillColor(C_BLUE2)
    canvas.rect(0, 0, sidebar_w, page_h * 0.68, stroke=0, fill=1)
    canvas.setFillColor(C_ACCENT)
    canvas.rect(0, page_h * 0.68, sidebar_w, page_h * 0.32, stroke=0, fill=1)
    canvas.setFillColor(colors.white)
    for y in (page_h * 0.14, page_h * 0.24, page_h * 0.48, page_h * 0.86):
        canvas.circle(sidebar_w / 2, y, 2.1, stroke=0, fill=1)
    canvas.restoreState()


def _draw_watermark(canvas, doc):
    page_w, page_h = doc.pagesize
    logo_path = _logo_path()
    canvas.saveState()
    canvas.setStrokeColor(C_BLUE2)
    _set_alpha(canvas, stroke=0.045)
    for radius in (72, 112, 154):
        canvas.circle(page_w / 2 + 45, page_h / 2 - 16, radius, stroke=1, fill=0)
    canvas.setFillColor(C_ACCENT)
    _set_alpha(canvas, fill=0.035, stroke=1)
    canvas.circle(page_w / 2 + 45, page_h / 2 - 16, 46, stroke=0, fill=1)
    if logo_path.exists():
        _set_alpha(canvas, fill=0.08)
        canvas.drawImage(str(logo_path), page_w / 2 - 56, page_h / 2 - 38, width=114, height=76, mask='auto')
    canvas.restoreState()


def _draw_header(canvas, doc, title, subtitle, report_label):
    page_w, page_h = doc.pagesize
    header_h = 4.5 * cm
    y0 = page_h - header_h
    logo_path = _logo_path()
    canvas.saveState()
    canvas.setFillColor(C_NAVY)
    canvas.rect(0, y0, page_w, header_h, stroke=0, fill=1)

    path = canvas.beginPath()
    path.moveTo(page_w * 0.66, y0)
    path.lineTo(page_w * 0.82, page_h)
    path.lineTo(page_w, page_h)
    path.lineTo(page_w, y0)
    path.close()
    canvas.setFillColor(C_BLUE1)
    canvas.drawPath(path, stroke=0, fill=1)

    path = canvas.beginPath()
    path.moveTo(page_w * 0.78, y0)
    path.lineTo(page_w * 0.92, page_h)
    path.lineTo(page_w, page_h)
    path.lineTo(page_w, y0)
    path.close()
    _set_alpha(canvas, fill=0.86)
    canvas.setFillColor(C_ACCENT)
    canvas.drawPath(path, stroke=0, fill=1)
    _set_alpha(canvas, fill=1)

    canvas.setFillColor(C_BLUE3)
    canvas.rect(0, y0 - 3, page_w, 3, stroke=0, fill=1)

    if logo_path.exists():
        canvas.drawImage(str(logo_path), 27 * mm, page_h - 2.08 * cm, width=24 * mm, height=16 * mm, mask='auto')
    canvas.setFillColor(colors.white)
    canvas.setFont('Helvetica-Bold', 17)
    canvas.drawString(27 * mm, page_h - 2.95 * cm, title)
    canvas.setFillColor(colors.HexColor('#BFDBFE'))
    canvas.setFont('Helvetica', 6.4)
    canvas.drawString(27 * mm, page_h - 3.35 * cm, subtitle)

    canvas.setFillColor(C_NAVY)
    canvas.setFont('Helvetica-Bold', 6.2)
    canvas.drawRightString(page_w - 23 * mm, page_h - 2.15 * cm, 'SISTEMA DE GESTION')
    canvas.setFont('Helvetica', 5.6)
    canvas.drawRightString(page_w - 23 * mm, page_h - 2.56 * cm, f'{report_label} registrados')
    canvas.restoreState()


def _draw_footer(canvas, doc):
    page_w, _page_h = doc.pagesize
    footer_h = 1.6 * cm
    canvas.saveState()
    canvas.setFillColor(C_NAVY)
    canvas.rect(0, 0, page_w, footer_h, stroke=0, fill=1)
    canvas.setFillColor(C_BLUE3)
    canvas.rect(0, footer_h, page_w, 2, stroke=0, fill=1)
    canvas.setFillColor(colors.HexColor('#BFDBFE'))
    canvas.setFont('Helvetica', 6)
    canvas.drawString(27 * mm, 15, '© 2026 GesPro - Sistema de Gestión Profesional')
    canvas.setFillColor(C_ACCENT)
    canvas.setFont('Helvetica-Bold', 7)
    canvas.drawRightString(page_w - 15 * mm, 15, f'Página {doc.page}')
    canvas.restoreState()


def _draw_report_page(title, generated_at):
    def _draw(canvas, doc):
        page_w, page_h = doc.pagesize
        canvas.saveState()
        canvas.setFillColor(C_BG)
        canvas.rect(0, 0, page_w, page_h, stroke=0, fill=1)
        canvas.restoreState()
        _draw_watermark(canvas, doc)
        _draw_sidebar(canvas, doc)
        _draw_header(canvas, doc, f'Reporte de {title}', f'Generado: {generated_at}', title)
        _draw_footer(canvas, doc)
    return _draw


def _matches_search(queryset, report_name: str, search: str):
    if not search:
        return queryset

    if report_name == 'usuarios':
        query = Q(nombre__icontains=search) | Q(apellido__icontains=search)
        if search.isdigit():
            query |= Q(numero_documento__icontains=search)
        if '@' in search:
            query |= Q(correo__icontains=search)
        return queryset.filter(query)

    if report_name in {'aprendices', 'instructores'}:
        query = Q(usuario__nombre__icontains=search) | Q(usuario__apellido__icontains=search)
        if search.isdigit():
            query |= Q(usuario__numero_documento__icontains=search)
        if '@' in search:
            query |= Q(usuario__correo__icontains=search)
        return queryset.filter(query)

    if report_name == 'fichas':
        return queryset.filter(codigo_ficha__icontains=search) | queryset.filter(programa_formacion__icontains=search)

    if report_name == 'trimestres':
        return queryset.filter(ficha__codigo_ficha__icontains=search)

    return queryset.filter(nombre__icontains=search)


def _total_trimestres_por_nivel(nivel: str | None) -> int:
    nivel_normalizado = (nivel or '').strip().lower()
    if nivel_normalizado == 'tecnico':
        return 4
    if nivel_normalizado == 'tecnologo':
        return 7
    return 0


def build_rows(report_name: str, filters: dict | None = None):
    filters = filters or {}
    search = (filters.get('q') or '').strip()

    if report_name == 'usuarios':
        queryset = Usuario.objects.select_related('rol').all()
        queryset = _matches_search(queryset, report_name, search)
        if filters.get('rol'):
            queryset = queryset.filter(rol__nombre_rol__iexact=filters['rol'])
        if filters.get('estado'):
            queryset = queryset.filter(estado=filters['estado'])
        headers = ['ID', 'Nombre', 'Apellido', 'Correo', 'Rol', 'Estado']
        rows = [
            [u.id, u.nombre, u.apellido, u.correo, u.rol.nombre_rol if u.rol else '', u.estado]
            for u in queryset.distinct()
        ]
        return headers, rows

    if report_name == 'aprendices':
        queryset = Aprendiz.objects.select_related('usuario', 'usuario__rol', 'ficha').filter(
            usuario__rol__nombre_rol__iexact='aprendiz',
        )
        queryset = _matches_search(queryset, report_name, search)
        if filters.get('ficha'):
            queryset = queryset.filter(ficha_id=filters['ficha'])
        headers = ['ID', 'Aprendiz', 'Tipo documento', 'Numero documento', 'Ficha', 'Rol', 'Correo']
        rows = [
            [
                a.id,
                f'{a.usuario.nombre} {a.usuario.apellido}'.strip(),
                a.usuario.tipo_documento,
                a.usuario.numero_documento,
                a.ficha.codigo_ficha if a.ficha else '',
                a.usuario.rol.nombre_rol if a.usuario.rol else 'aprendiz',
                a.usuario.correo,
            ]
            for a in queryset.distinct()
        ]
        return headers, rows

    if report_name == 'instructores':
        queryset = Instructor.objects.select_related('usuario', 'usuario__rol', 'ficha').filter(
            usuario__rol__nombre_rol__iexact='instructor',
        )
        queryset = _matches_search(queryset, report_name, search)
        if filters.get('ficha'):
            queryset = queryset.filter(ficha_id=filters['ficha'])
        headers = ['ID', 'Instructor', 'Ficha', 'Especialidad', 'Rol', 'Correo']
        rows = [
            [
                i.id,
                f'{i.usuario.nombre} {i.usuario.apellido}'.strip(),
                i.ficha.codigo_ficha if i.ficha else '',
                i.especialidad or '',
                i.usuario.rol.nombre_rol if i.usuario.rol else 'instructor',
                i.usuario.correo,
            ]
            for i in queryset.distinct()
        ]
        return headers, rows

    if report_name == 'fichas':
        queryset = Ficha.objects.all()
        queryset = _matches_search(queryset, report_name, search)
        if filters.get('estado'):
            queryset = queryset.filter(estado=filters['estado'])
        headers = ['ID', 'Código', 'Programa', 'Nivel', 'Jornada', 'Modalidad', 'Inicio', 'Fin', 'Estado']
        rows = [
            [f.id, f.codigo_ficha, f.programa_formacion, f.nivel, f.jornada, f.modalidad, f.fecha_inicio, f.fecha_fin, f.estado]
            for f in queryset.distinct()
        ]
        return headers, rows

    if report_name == 'trimestres':
        queryset = Ficha.objects.all()
        if search:
            queryset = queryset.filter(codigo_ficha__icontains=search) | queryset.filter(programa_formacion__icontains=search)
        if filters.get('ficha'):
            queryset = queryset.filter(id=filters['ficha'])
        if filters.get('estado'):
            queryset = queryset.filter(estado=filters['estado'])
        headers = ['ID', 'Ficha', 'Programa', 'Nivel', 'Total trimestres', 'Estado', 'Inicio', 'Fin']
        rows = []
        for ficha in queryset.distinct():
            nivel = ficha.nivel
            total_trimestres = _total_trimestres_por_nivel(nivel)
            rows.append(
                [
                    ficha.id,
                    ficha.codigo_ficha,
                    ficha.programa_formacion,
                    'Tecnólogo' if nivel == Ficha.Nivel.TECNOLOGO else 'Técnico',
                    f'{total_trimestres} trimestres' if total_trimestres else '',
                    ficha.estado,
                    ficha.fecha_inicio,
                    ficha.fecha_fin,
                ]
            )
        return headers, rows

    return build_rows('usuarios', filters)


def excel_response(filename: str, sheet_name: str, headers: list[str], rows: list[list]):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31]
    worksheet.append([f'Reporte de {sheet_name}'])
    worksheet.append([f'Generado: {timezone.localtime().strftime("%d/%m/%Y %I:%M %p")}'])
    worksheet.append([])
    worksheet.append(headers)

    header_fill = PatternFill(fill_type='solid', start_color='1D4ED8', end_color='1D4ED8')
    header_font = Font(color='FFFFFF', bold=True)

    worksheet['A1'].font = Font(bold=True, size=16, color='1F2937')
    worksheet['A2'].font = Font(italic=True, color='64748B')

    for cell in worksheet[4]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        worksheet.append(row)

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = '' if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 4, 40)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _column_widths(headers: list[str], available_width: float) -> list[float]:
    width_map = {
        'id': 0.8 * cm,
        'rol': 1.7 * cm,
        'estado': 2.0 * cm,
        'ficha': 1.6 * cm,
        'numero': 1.5 * cm,
        'trimestre': 1.9 * cm,
        'total trimestres': 2.5 * cm,
        'nivel': 2.0 * cm,
        'tipo documento': 2.4 * cm,
        'numero documento': 2.7 * cm,
        'documento': 2.7 * cm,
        'correo': 4.5 * cm,
        'aprendiz': 3.2 * cm,
        'instructor': 3.2 * cm,
        'nombre': 2.4 * cm,
        'apellido': 2.4 * cm,
        'codigo': 1.7 * cm,
        'programa': 3.4 * cm,
        'especialidad': 3.0 * cm,
        'fecha': 2.1 * cm,
        'inicio': 2.1 * cm,
        'fin': 2.1 * cm,
    }
    widths = [width_map.get(str(header).lower(), 2.2 * cm) for header in headers]
    total = sum(widths)
    if total > available_width:
        factor = available_width / total
        widths = [width * factor for width in widths]
    elif total < available_width:
        extra = (available_width - total) / len(widths)
        widths = [width + extra for width in widths]
    return widths


def pdf_response(filename: str, title: str, headers: list[str], rows: list[list]):
    output = BytesIO()
    page_size = landscape(letter)
    document = SimpleDocTemplate(
        output,
        pagesize=page_size,
        leftMargin=27 * mm,
        rightMargin=14 * mm,
        topMargin=5.25 * cm,
        bottomMargin=2.15 * cm,
    )
    generated_at = timezone.localtime().strftime('%d/%m/%Y %I:%M %p')

    table_rows = [headers]
    cell_styles = []
    id_index = next((idx for idx, header in enumerate(headers) if header.lower() == 'id'), None)
    role_index = next((idx for idx, header in enumerate(headers) if header.lower() == 'rol'), None)
    state_index = next((idx for idx, header in enumerate(headers) if header.lower() == 'estado'), None)

    for row_number, row in enumerate(rows, start=1):
        rendered_row = []
        for idx, cell in enumerate(row):
            raw_value = '' if cell is None else str(cell)
            lower_value = raw_value.lower()
            display_value = raw_value
            if idx == id_index:
                cell_styles.extend(
                    [
                        ('TEXTCOLOR', (idx, row_number), (idx, row_number), C_BLUE2),
                        ('FONTNAME', (idx, row_number), (idx, row_number), 'Helvetica-Bold'),
                    ]
                )
            elif idx == role_index:
                role_color = {
                    'aprendiz': C_BLUE2,
                    'instructor': C_VIOLET,
                    'administrador': C_AMBER,
                    'admin': C_AMBER,
                }.get(lower_value, C_BLUE2)
                cell_styles.extend(
                    [
                        ('TEXTCOLOR', (idx, row_number), (idx, row_number), role_color),
                        ('FONTNAME', (idx, row_number), (idx, row_number), 'Helvetica-Bold'),
                    ]
                )
            elif idx == state_index:
                if lower_value == 'activo':
                    display_value = 'ACTIVO'
                    state_color = C_ACTIVO
                elif lower_value == 'inactivo':
                    display_value = 'INACTIVO'
                    state_color = C_INACTIVO
                elif lower_value == 'activa':
                    display_value = 'ACTIVA'
                    state_color = C_ACTIVO
                elif lower_value == 'inactiva':
                    display_value = 'INACTIVA'
                    state_color = C_INACTIVO
                elif lower_value == 'aprobado':
                    display_value = 'APROBADO'
                    state_color = C_ACTIVO
                elif lower_value in {'no aprobado', 'no_aprobado'}:
                    display_value = 'NO APROBADO'
                    state_color = C_INACTIVO
                elif lower_value in {'pendiente', 'en proceso', 'en_proceso'}:
                    display_value = 'EN PROCESO'
                    state_color = C_AMBER
                else:
                    display_value = raw_value.upper()
                    state_color = C_AMBER
                cell_styles.extend(
                    [
                        ('TEXTCOLOR', (idx, row_number), (idx, row_number), state_color),
                        ('FONTNAME', (idx, row_number), (idx, row_number), 'Helvetica-Bold'),
                    ]
                )
            rendered_row.append(display_value)
        table_rows.append(rendered_row)

    available_width = page_size[0] - document.leftMargin - document.rightMargin
    col_widths = _column_widths(headers, available_width)

    table = Table(table_rows, repeatRows=1, colWidths=col_widths)
    row_backgrounds = []
    for row_index in range(1, len(table_rows)):
        row_backgrounds.append(('BACKGROUND', (0, row_index), (-1, row_index), C_ROW_ALT if row_index % 2 == 0 else colors.white))
    table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), C_HEADER_ROW),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('FONTSIZE', (0, 1), (-1, -1), 6.2),
                ('GRID', (0, 0), (-1, -1), 0.35, C_BORDER),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('WORDWRAP', (0, 0), (-1, -1), 'CJK'),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
                ('TOPPADDING', (0, 0), (-1, 0), 5),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
            ] + row_backgrounds + cell_styles
        )
    )

    state_values = [str(row[state_index]).lower() for row in rows] if state_index is not None else []
    is_final_report = 'evaluacion final' in title.lower() or 'evaluación final' in title.lower()
    if is_final_report:
        approved_count = sum(1 for value in state_values if value == 'aprobado')
        not_approved_count = sum(1 for value in state_values if value in {'no aprobado', 'no_aprobado'})
        pending_count = sum(1 for value in state_values if value in {'pendiente', 'en proceso', 'en_proceso'})
        stat_cards = [
            StatCard(len(rows), 'Aprendices', C_ACCENT),
            StatCard(approved_count, 'Aprobados', C_ACTIVO),
            StatCard(not_approved_count, 'No aprobados', C_INACTIVO),
            StatCard(pending_count, 'En proceso', C_AMBER),
        ]
    else:
        active_count = state_values.count('activo')
        inactive_count = state_values.count('inactivo')
        role_count = len({str(row[role_index]).lower() for row in rows if str(row[role_index]).strip()}) if role_index is not None else len(headers)
        stat_cards = [
            StatCard(len(rows), f'Total {title}', C_ACCENT),
            StatCard(active_count, 'Activos', C_ACTIVO),
            StatCard(inactive_count, 'Inactivos', C_INACTIVO),
            StatCard(role_count, 'Roles' if role_index is not None else 'Campos', C_AMBER),
        ]
    stats_table = Table([stat_cards], colWidths=[3.9 * cm, 3.9 * cm, 3.9 * cm, 3.9 * cm], hAlign='CENTER')
    stats_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))

    story = []
    story.extend([
        SectionBar('Resumen estadístico', width=available_width),
        Spacer(1, 8),
        stats_table,
        Spacer(1, 14),
        SectionBar(f'Listado de {title} registrados', width=available_width),
        Spacer(1, 8),
        table,
        Spacer(1, 8),
        AccentLine(width=available_width),
    ])
    document.build(story, onFirstPage=_draw_report_page(title, generated_at), onLaterPages=_draw_report_page(title, generated_at))
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
